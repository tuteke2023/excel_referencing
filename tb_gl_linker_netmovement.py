#!/usr/bin/env python3
"""
TB-GL Linker (Net Movement Version): Links TB accounts to Net Movement figures in GL
Instead of linking to account headers (A9829), this version links to the actual
Net Movement cell (e.g., F9846) and displays the value as the hyperlink text.
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
import os
import re
from difflib import SequenceMatcher
import argparse


class TBGLLinkerNetMovement:
    def __init__(self, tb_file, gl_file, output_file=None):
        self.tb_file = tb_file
        self.gl_file = gl_file
        self.output_file = output_file or tb_file.replace('.xlsx', '_linked.xlsx')
        self.tb_wb = None
        self.gl_wb = None
        self.tb_sheet = None
        self.gl_sheet = None
        self.account_mappings = {}
        self.tb_config = {}
        self.gl_accounts = {}  # {account_name: {header_row, net_movement_row, net_movement_col, target_cell, value}}
        self.gl_debit_col = None
        self.gl_credit_col = None

    def load_workbooks(self):
        """Load both workbooks and identify the main sheets."""
        print(f"Loading {self.tb_file}...")
        self.tb_wb = load_workbook(self.tb_file)
        print(f"Loading {self.gl_file}...")
        self.gl_wb = load_workbook(self.gl_file)

        # Find TB sheet (usually first sheet or one with 'TB' in name)
        tb_sheet_name = self._find_sheet(self.tb_wb, ['TB', 'Trial Balance', 'Trial_Balance'])
        self.tb_sheet = self.tb_wb[tb_sheet_name]
        print(f"Using TB sheet: {tb_sheet_name}")

        # Find GL sheet
        gl_sheet_name = self._find_sheet(self.gl_wb, ['GL', 'General Ledger', 'General_Ledger'])
        self.gl_sheet = self.gl_wb[gl_sheet_name]
        print(f"Using GL sheet: {gl_sheet_name}")

    def _find_sheet(self, workbook, keywords):
        """Find sheet by keywords in sheet names."""
        sheet_names = workbook.sheetnames

        # First try exact match
        for sheet in sheet_names:
            for keyword in keywords:
                if keyword.lower() in sheet.lower():
                    return sheet

        # If no match found, return first sheet
        return sheet_names[0]

    def analyze_tb_structure(self):
        """Analyze TB structure to find account columns."""
        print("\nAnalyzing Trial Balance structure...")

        # Find header row (usually contains 'Account', 'Debit', 'Credit')
        header_row = None
        account_col = None
        account_name_col = None
        debit_col = None
        credit_col = None

        for row in range(1, min(20, self.tb_sheet.max_row + 1)):
            row_values = []
            for col in range(1, min(20, self.tb_sheet.max_column + 1)):
                cell_value = self.tb_sheet.cell(row, col).value
                if cell_value:
                    row_values.append((col, str(cell_value).lower()))

            # Look for key headers
            for col, value in row_values:
                if any(keyword in value for keyword in ['account', 'acct', 'code']):
                    if 'name' in value or 'description' in value:
                        account_name_col = col
                    else:
                        account_col = col
                    header_row = row
                elif 'debit' in value:
                    debit_col = col
                elif 'credit' in value:
                    credit_col = col

            if header_row and (account_col or account_name_col) and (debit_col or credit_col):
                break

        if not header_row:
            raise ValueError("Could not find header row in Trial Balance")

        self.tb_config = {
            'header_row': header_row,
            'account_col': account_col,
            'account_name_col': account_name_col,
            'debit_col': debit_col,
            'credit_col': credit_col,
            'data_start_row': header_row + 1
        }

        print(f"TB Structure found:")
        print(f"  Header row: {header_row}")
        print(f"  Account code column: {get_column_letter(account_col) if account_col else 'Not found'}")
        print(f"  Account name column: {get_column_letter(account_name_col) if account_name_col else 'Not found'}")
        print(f"  Debit column: {get_column_letter(debit_col) if debit_col else 'Not found'}")
        print(f"  Credit column: {get_column_letter(credit_col) if credit_col else 'Not found'}")

    def _find_gl_debit_credit_cols(self, near_row=None):
        """Detect Debit/Credit column positions in GL sheet."""
        if self.gl_debit_col and self.gl_credit_col:
            return self.gl_debit_col, self.gl_credit_col

        debit_col = None
        credit_col = None

        # Search in header area first (rows 1-20)
        search_rows = list(range(1, min(21, self.gl_sheet.max_row + 1)))

        # Also search near the specified row if provided
        if near_row:
            search_rows.extend(range(max(1, near_row - 5), min(near_row + 10, self.gl_sheet.max_row + 1)))

        for row in search_rows:
            for col in range(1, min(20, self.gl_sheet.max_column + 1)):
                cell_value = self.gl_sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower().strip()
                    if 'debit' in cell_lower and not debit_col:
                        debit_col = col
                    elif 'credit' in cell_lower and not credit_col:
                        credit_col = col

            if debit_col and credit_col:
                break

        # Fallback defaults if not found (common positions)
        if not debit_col:
            debit_col = 5  # Column E
        if not credit_col:
            credit_col = 6  # Column F

        self.gl_debit_col = debit_col
        self.gl_credit_col = credit_col

        print(f"  GL Debit column: {get_column_letter(debit_col)}, Credit column: {get_column_letter(credit_col)}")
        return debit_col, credit_col

    def _get_nonzero_column(self, row, debit_col, credit_col):
        """Get the column with a non-zero value. Returns (column_number, value)."""
        debit_value = self.gl_sheet.cell(row, debit_col).value
        credit_value = self.gl_sheet.cell(row, credit_col).value

        # Convert to numbers, handling None and strings
        try:
            debit_num = float(debit_value) if debit_value else 0
        except (ValueError, TypeError):
            debit_num = 0

        try:
            credit_num = float(credit_value) if credit_value else 0
        except (ValueError, TypeError):
            credit_num = 0

        # If both have values, use Credit (net typically shows there)
        if credit_num != 0:
            return credit_col, credit_num
        elif debit_num != 0:
            return debit_col, debit_num
        else:
            # Both are 0, default to Debit column
            return debit_col, 0

    def _find_net_movement(self, header_row, next_account_row=None):
        """
        Search rows below account header for "Net Movement" text.
        Returns dict with net_movement info or None if not found.
        """
        debit_col, credit_col = self._find_gl_debit_credit_cols(header_row)

        # Determine search limit (stop at next account or reasonable limit)
        max_search = next_account_row if next_account_row else header_row + 500
        max_search = min(max_search, self.gl_sheet.max_row + 1)

        for row in range(header_row + 1, max_search):
            # Check first few columns for "Net Movement" text
            for col in range(1, min(5, self.gl_sheet.max_column + 1)):
                cell_value = self.gl_sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    if 'net movement' in cell_value.lower():
                        # Found Net Movement row - get the non-zero column
                        target_col, value = self._get_nonzero_column(row, debit_col, credit_col)
                        target_cell = f"{get_column_letter(target_col)}{row}"

                        return {
                            'net_movement_row': row,
                            'net_movement_col': target_col,
                            'target_cell': target_cell,
                            'value': value
                        }

        return None

    def _is_account_header(self, row, value):
        """Determine if a cell contains an account header."""
        # Check if next rows have date/transaction data
        if row >= self.gl_sheet.max_row:
            return False

        # Account headers usually don't have dates or numbers in the same row
        for col in range(2, min(10, self.gl_sheet.max_column + 1)):
            cell_value = self.gl_sheet.cell(row, col).value
            if cell_value:
                return False

        # Check if following rows have transaction-like data
        next_row_has_data = False
        for check_row in range(row + 1, min(row + 5, self.gl_sheet.max_row + 1)):
            cell_value = self.gl_sheet.cell(check_row, 1).value
            if cell_value and (isinstance(cell_value, (int, float)) or
                             self._is_date_like(str(cell_value))):
                next_row_has_data = True
                break

        return next_row_has_data

    def _is_date_like(self, value):
        """Check if string looks like a date."""
        date_patterns = [r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', r'\d{4}[/-]\d{1,2}[/-]\d{1,2}']
        return any(re.match(pattern, value) for pattern in date_patterns)

    def analyze_gl_structure(self):
        """Analyze GL structure to find account sections and their Net Movement rows."""
        print("\nAnalyzing General Ledger structure...")

        # First pass: find all account headers and their row numbers
        account_headers = []

        for row in range(1, self.gl_sheet.max_row + 1):
            # Check first few columns for account names
            for col in range(1, min(5, self.gl_sheet.max_column + 1)):
                cell_value = self.gl_sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    # Check if this looks like an account header
                    if col == 1 and self._is_account_header(row, cell_value):
                        account_name = cell_value.strip()
                        account_headers.append((account_name, row))

        print(f"Found {len(account_headers)} GL account sections")

        # Detect Debit/Credit columns
        self._find_gl_debit_credit_cols()

        # Second pass: find Net Movement for each account
        print("\nFinding Net Movement rows...")
        for i, (account_name, header_row) in enumerate(account_headers):
            # Determine the next account's row (to limit search)
            next_account_row = account_headers[i + 1][1] if i + 1 < len(account_headers) else None

            # Find Net Movement row
            net_movement_info = self._find_net_movement(header_row, next_account_row)

            if net_movement_info:
                self.gl_accounts[account_name] = {
                    'header_row': header_row,
                    'net_movement_row': net_movement_info['net_movement_row'],
                    'net_movement_col': net_movement_info['net_movement_col'],
                    'target_cell': net_movement_info['target_cell'],
                    'value': net_movement_info['value']
                }
                print(f"  {account_name}: Net Movement at {net_movement_info['target_cell']} = {net_movement_info['value']}")
            else:
                # Fallback to account header
                self.gl_accounts[account_name] = {
                    'header_row': header_row,
                    'net_movement_row': header_row,
                    'net_movement_col': 1,
                    'target_cell': f"A{header_row}",
                    'value': None
                }
                print(f"  {account_name}: No Net Movement found, using header A{header_row}")

        print(f"\nProcessed {len(self.gl_accounts)} accounts with Net Movement lookup")

    def match_accounts(self):
        """Match TB accounts to GL accounts using fuzzy matching."""
        print("\nMatching accounts between TB and GL...")

        # Get TB accounts
        tb_accounts = []
        for row in range(self.tb_config['data_start_row'], self.tb_sheet.max_row + 1):
            account_name = None

            # Try to get account name from name column first
            if self.tb_config['account_name_col']:
                account_name = self.tb_sheet.cell(row, self.tb_config['account_name_col']).value

            # If no name column or empty, try to find account name in other columns
            if not account_name:
                # Check columns near the account code column for account names
                for col_offset in [1, -1, 2, -2]:  # Check adjacent columns
                    check_col = (self.tb_config['account_col'] or 1) + col_offset
                    if 1 <= check_col <= self.tb_sheet.max_column:
                        cell_value = self.tb_sheet.cell(row, check_col).value
                        if cell_value and isinstance(cell_value, str) and len(cell_value) > 3:
                            # Check if it looks like an account name (not a number or code)
                            if not cell_value.replace('.', '').replace('-', '').isdigit():
                                account_name = cell_value
                                if not self.tb_config['account_name_col']:
                                    self.tb_config['account_name_col'] = check_col
                                break

            if account_name and isinstance(account_name, str):
                tb_accounts.append((row, account_name.strip()))

        # Match each TB account to GL account
        matched = 0
        for tb_row, tb_account in tb_accounts:
            best_match = None
            best_score = 0

            for gl_account, gl_info in self.gl_accounts.items():
                # Calculate similarity score
                score = SequenceMatcher(None, tb_account.lower(), gl_account.lower()).ratio()

                # Bonus for exact match
                if tb_account.lower() == gl_account.lower():
                    score = 1.0

                if score > best_score and score > 0.8:  # 80% similarity threshold
                    best_score = score
                    best_match = (gl_account, gl_info)

            if best_match:
                self.account_mappings[tb_row] = best_match
                matched += 1
                gl_name, gl_info = best_match
                print(f"  Matched: '{tb_account}' -> '{gl_name}' @ {gl_info['target_cell']} (score: {best_score:.2f})")
            else:
                print(f"  No match found for: '{tb_account}'")

        if len(tb_accounts) > 0:
            print(f"\nMatched {matched} out of {len(tb_accounts)} accounts ({matched/len(tb_accounts)*100:.1f}%)")
        else:
            print("\nNo accounts found in Trial Balance")

    def copy_gl_sheet(self):
        """Copy GL sheet to TB workbook."""
        print("\nCopying General Ledger sheet to Trial Balance workbook...")

        # Create new sheet in TB workbook
        gl_sheet_name = 'General Ledger Detail'
        if gl_sheet_name in self.tb_wb.sheetnames:
            del self.tb_wb[gl_sheet_name]

        new_sheet = self.tb_wb.create_sheet(title=gl_sheet_name)

        # Copy all data
        for row in self.gl_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                # Copy formatting
                if cell.has_style:
                    new_cell.font = openpyxl.styles.Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color
                    )
                    new_cell.fill = openpyxl.styles.PatternFill(
                        fill_type=cell.fill.fill_type,
                        fgColor=cell.fill.fgColor,
                        bgColor=cell.fill.bgColor
                    )
                    new_cell.alignment = openpyxl.styles.Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text
                    )
                    new_cell.number_format = cell.number_format

        # Copy column widths
        for column_cells in self.gl_sheet.columns:
            column_letter = column_cells[0].column_letter
            if self.gl_sheet.column_dimensions[column_letter].width:
                new_sheet.column_dimensions[column_letter].width = self.gl_sheet.column_dimensions[column_letter].width

        print("GL sheet copied successfully!")

    def add_hyperlinks(self):
        """Add hyperlinks to TB sheet pointing to Net Movement cells."""
        print("\nAdding hyperlinks to Trial Balance (Net Movement version)...")

        # Find the best column for hyperlinks (after Credit or last column)
        if self.tb_config['credit_col']:
            hyperlink_col = self.tb_config['credit_col'] + 1
        elif self.tb_config['debit_col']:
            hyperlink_col = self.tb_config['debit_col'] + 1
        else:
            hyperlink_col = self.tb_sheet.max_column + 1

        # Add header - changed to "GL Reference"
        self.tb_sheet.cell(self.tb_config['header_row'], hyperlink_col, 'GL Reference')

        # Add hyperlinks for matched accounts
        for tb_row, (gl_account, gl_info) in self.account_mappings.items():
            cell = self.tb_sheet.cell(tb_row, hyperlink_col)
            target_cell = gl_info['target_cell']
            value = gl_info['value']

            # Format the value for display
            if value is not None and value != 0:
                # Format number with commas, no decimal if whole number
                if isinstance(value, float) and value == int(value):
                    display_value = f"{int(value):,}"
                else:
                    display_value = f"{value:,.2f}" if isinstance(value, float) else str(value)
            else:
                display_value = "0"

            # Create hyperlink formula with value as display text
            formula = f'=HYPERLINK("#\'General Ledger Detail\'!{target_cell}", "{display_value}")'
            cell.value = formula

        # Add "N/A" for unmatched accounts
        for row in range(self.tb_config['data_start_row'], self.tb_sheet.max_row + 1):
            if row not in self.account_mappings:
                account_name = self.tb_sheet.cell(row, self.tb_config['account_name_col']).value
                if account_name:
                    self.tb_sheet.cell(row, hyperlink_col, 'N/A')

        print(f"Hyperlinks added in column {get_column_letter(hyperlink_col)}")

    def save_workbook(self):
        """Save the modified workbook."""
        print(f"\nSaving linked workbook as: {self.output_file}")
        self.tb_wb.save(self.output_file)
        print("Done! The linked workbook has been created successfully.")

    def run(self):
        """Execute the complete linking process."""
        try:
            self.load_workbooks()
            self.analyze_tb_structure()
            self.analyze_gl_structure()
            self.match_accounts()
            self.copy_gl_sheet()
            self.add_hyperlinks()
            self.save_workbook()

            print(f"\n{'='*60}")
            print(f"SUCCESS! Your linked file is ready: {self.output_file}")
            print(f"{'='*60}")
            print(f"   - {len(self.account_mappings)} accounts linked to Net Movement figures")
            print(f"   - GL data copied as 'General Ledger Detail' sheet")
            print(f"   - Click the value links to navigate to GL Net Movement cells")
            print(f"{'='*60}")

        except Exception as e:
            print(f"\nError: {str(e)}")
            sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description='Link Trial Balance accounts to General Ledger Net Movement figures',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s TB.xlsx GL.xlsx
  %(prog)s trial_balance.xlsx general_ledger.xlsx -o combined.xlsx

This version links to Net Movement figures (e.g., F9846) instead of account headers.
The hyperlink text shows the actual Net Movement value for quick reference.
        """
    )

    parser.add_argument('tb_file', help='Trial Balance Excel file')
    parser.add_argument('gl_file', help='General Ledger Excel file')
    parser.add_argument('-o', '--output', help='Output file name (default: TB_file_linked.xlsx)')

    args = parser.parse_args()

    # Validate input files exist
    if not os.path.exists(args.tb_file):
        print(f"Error: TB file '{args.tb_file}' not found")
        sys.exit(1)

    if not os.path.exists(args.gl_file):
        print(f"Error: GL file '{args.gl_file}' not found")
        sys.exit(1)

    # Create linker and run
    linker = TBGLLinkerNetMovement(args.tb_file, args.gl_file, args.output)
    linker.run()


if __name__ == '__main__':
    main()
