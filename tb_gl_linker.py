#!/usr/bin/env python3
"""
TB-GL Linker: Automatically link Trial Balance accounts to General Ledger details
This script copies GL data into TB file and creates hyperlinks for easy navigation.
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
import os
import re
from difflib import SequenceMatcher
import argparse


class TBGLLinker:
    def __init__(self, tb_file, gl_file, output_file=None):
        self.tb_file = tb_file
        self.gl_file = gl_file
        self.output_file = output_file or tb_file.replace('.xlsx', '_linked.xlsx')
        self.tb_wb = None
        self.gl_wb = None
        self.tb_sheet = None
        self.gl_sheet = None
        self.account_mappings = {}
        
    def load_workbooks(self):
        """Load both workbooks and identify the main sheets."""
        print(f"Loading {self.tb_file}...")
        self.tb_wb = load_workbook(self.tb_file)
        print(f"Loading {self.gl_file}...")
        self.gl_wb = load_workbook(self.gl_file)
        
        # Find TB sheet (usually first sheet or one with 'TB' in name)
        tb_sheet_name = self._find_sheet(self.tb_wb, ['TB', 'Trial Balance', 'Trial_Balance'])
        self.tb_sheet = self.tb_wb[tb_sheet_name]
        print(f"Using TB sheet: {tb_sheet_name}")
        
        # Find GL sheet
        gl_sheet_name = self._find_sheet(self.gl_wb, ['GL', 'General Ledger', 'General_Ledger'])
        self.gl_sheet = self.gl_wb[gl_sheet_name]
        print(f"Using GL sheet: {gl_sheet_name}")
        
    def _find_sheet(self, workbook, keywords):
        """Find sheet by keywords in sheet names."""
        sheet_names = workbook.sheetnames
        
        # First try exact match
        for sheet in sheet_names:
            for keyword in keywords:
                if keyword.lower() in sheet.lower():
                    return sheet
        
        # If no match found, return first sheet
        return sheet_names[0]
    
    def analyze_tb_structure(self):
        """Analyze TB structure to find account columns."""
        print("\nAnalyzing Trial Balance structure...")
        
        # Find header row (usually contains 'Account', 'Debit', 'Credit')
        header_row = None
        account_col = None
        account_name_col = None
        debit_col = None
        credit_col = None
        
        for row in range(1, min(20, self.tb_sheet.max_row + 1)):
            row_values = []
            for col in range(1, min(20, self.tb_sheet.max_column + 1)):
                cell_value = self.tb_sheet.cell(row, col).value
                if cell_value:
                    row_values.append((col, str(cell_value).lower()))
            
            # Look for key headers
            for col, value in row_values:
                if any(keyword in value for keyword in ['account', 'acct', 'code']):
                    if 'name' in value or 'description' in value:
                        account_name_col = col
                    else:
                        account_col = col
                    header_row = row
                elif 'debit' in value:
                    debit_col = col
                elif 'credit' in value:
                    credit_col = col
            
            if header_row and (account_col or account_name_col) and (debit_col or credit_col):
                break
        
        if not header_row:
            raise ValueError("Could not find header row in Trial Balance")
        
        self.tb_config = {
            'header_row': header_row,
            'account_col': account_col,
            'account_name_col': account_name_col,
            'debit_col': debit_col,
            'credit_col': credit_col,
            'data_start_row': header_row + 1
        }
        
        print(f"TB Structure found:")
        print(f"  Header row: {header_row}")
        print(f"  Account code column: {get_column_letter(account_col) if account_col else 'Not found'}")
        print(f"  Account name column: {get_column_letter(account_name_col) if account_name_col else 'Not found'}")
        print(f"  Debit column: {get_column_letter(debit_col) if debit_col else 'Not found'}")
        print(f"  Credit column: {get_column_letter(credit_col) if credit_col else 'Not found'}")
        
    def analyze_gl_structure(self):
        """Analyze GL structure to find account sections."""
        print("\nAnalyzing General Ledger structure...")
        
        # GL typically has account names as section headers
        # Find all account sections
        self.gl_accounts = {}
        
        for row in range(1, self.gl_sheet.max_row + 1):
            # Check first few columns for account names
            for col in range(1, min(5, self.gl_sheet.max_column + 1)):
                cell_value = self.gl_sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    # Check if this looks like an account header
                    # Usually it's in column A and followed by transaction data
                    if col == 1 and self._is_account_header(row, cell_value):
                        account_name = cell_value.strip()
                        self.gl_accounts[account_name] = row
        
        print(f"Found {len(self.gl_accounts)} GL account sections")
        
    def _is_account_header(self, row, value):
        """Determine if a cell contains an account header."""
        # Check if next rows have date/transaction data
        if row >= self.gl_sheet.max_row:
            return False
        
        # Account headers usually don't have dates or numbers in the same row
        for col in range(2, min(10, self.gl_sheet.max_column + 1)):
            cell_value = self.gl_sheet.cell(row, col).value
            if cell_value:
                return False
        
        # Check if following rows have transaction-like data
        next_row_has_data = False
        for check_row in range(row + 1, min(row + 5, self.gl_sheet.max_row + 1)):
            cell_value = self.gl_sheet.cell(check_row, 1).value
            if cell_value and (isinstance(cell_value, (int, float)) or 
                             self._is_date_like(str(cell_value))):
                next_row_has_data = True
                break
        
        return next_row_has_data
    
    def _is_date_like(self, value):
        """Check if string looks like a date."""
        date_patterns = [r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', r'\d{4}[/-]\d{1,2}[/-]\d{1,2}']
        return any(re.match(pattern, value) for pattern in date_patterns)
    
    def match_accounts(self):
        """Match TB accounts to GL accounts using fuzzy matching."""
        print("\nMatching accounts between TB and GL...")
        
        # Get TB accounts
        tb_accounts = []
        for row in range(self.tb_config['data_start_row'], self.tb_sheet.max_row + 1):
            account_name = None
            
            # Try to get account name from name column first
            if self.tb_config['account_name_col']:
                account_name = self.tb_sheet.cell(row, self.tb_config['account_name_col']).value
            
            # If no name column or empty, try to find account name in other columns
            if not account_name:
                # Check columns near the account code column for account names
                for col_offset in [1, -1, 2, -2]:  # Check adjacent columns
                    check_col = (self.tb_config['account_col'] or 1) + col_offset
                    if 1 <= check_col <= self.tb_sheet.max_column:
                        cell_value = self.tb_sheet.cell(row, check_col).value
                        if cell_value and isinstance(cell_value, str) and len(cell_value) > 3:
                            # Check if it looks like an account name (not a number or code)
                            if not cell_value.replace('.', '').replace('-', '').isdigit():
                                account_name = cell_value
                                if not self.tb_config['account_name_col']:
                                    self.tb_config['account_name_col'] = check_col
                                break
            
            if account_name and isinstance(account_name, str):
                tb_accounts.append((row, account_name.strip()))
        
        # Match each TB account to GL account
        matched = 0
        for tb_row, tb_account in tb_accounts:
            best_match = None
            best_score = 0
            
            for gl_account, gl_row in self.gl_accounts.items():
                # Calculate similarity score
                score = SequenceMatcher(None, tb_account.lower(), gl_account.lower()).ratio()
                
                # Bonus for exact match
                if tb_account.lower() == gl_account.lower():
                    score = 1.0
                
                if score > best_score and score > 0.8:  # 80% similarity threshold
                    best_score = score
                    best_match = (gl_account, gl_row)
            
            if best_match:
                self.account_mappings[tb_row] = best_match
                matched += 1
                print(f"  Matched: '{tb_account}' -> '{best_match[0]}' (score: {best_score:.2f})")
            else:
                print(f"  No match found for: '{tb_account}'")
        
        if len(tb_accounts) > 0:
            print(f"\nMatched {matched} out of {len(tb_accounts)} accounts ({matched/len(tb_accounts)*100:.1f}%)")
        else:
            print("\nNo accounts found in Trial Balance")
        
    def copy_gl_sheet(self):
        """Copy GL sheet to TB workbook."""
        print("\nCopying General Ledger sheet to Trial Balance workbook...")
        
        # Create new sheet in TB workbook
        gl_sheet_name = 'General Ledger Detail'
        if gl_sheet_name in self.tb_wb.sheetnames:
            del self.tb_wb[gl_sheet_name]
        
        new_sheet = self.tb_wb.create_sheet(title=gl_sheet_name)
        
        # Copy all data
        for row in self.gl_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                # Copy formatting
                if cell.has_style:
                    new_cell.font = openpyxl.styles.Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color
                    )
                    new_cell.fill = openpyxl.styles.PatternFill(
                        fill_type=cell.fill.fill_type,
                        fgColor=cell.fill.fgColor,
                        bgColor=cell.fill.bgColor
                    )
                    new_cell.alignment = openpyxl.styles.Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text
                    )
                    new_cell.number_format = cell.number_format
        
        # Copy column widths
        for column_cells in self.gl_sheet.columns:
            column_letter = column_cells[0].column_letter
            if self.gl_sheet.column_dimensions[column_letter].width:
                new_sheet.column_dimensions[column_letter].width = self.gl_sheet.column_dimensions[column_letter].width
        
        print("GL sheet copied successfully!")
        
    def add_hyperlinks(self):
        """Add hyperlinks to TB sheet."""
        print("\nAdding hyperlinks to Trial Balance...")
        
        # Find the best column for hyperlinks (after Credit or last column)
        if self.tb_config['credit_col']:
            hyperlink_col = self.tb_config['credit_col'] + 1
        elif self.tb_config['debit_col']:
            hyperlink_col = self.tb_config['debit_col'] + 1
        else:
            hyperlink_col = self.tb_sheet.max_column + 1
        
        # Add header
        self.tb_sheet.cell(self.tb_config['header_row'], hyperlink_col, 'Reference')
        
        # Add hyperlinks for matched accounts
        for tb_row, (gl_account, gl_row) in self.account_mappings.items():
            cell = self.tb_sheet.cell(tb_row, hyperlink_col)
            formula = f'=HYPERLINK("#\'General Ledger Detail\'!A{gl_row}", "View Details")'
            cell.value = formula
        
        # Add "N/A" for unmatched accounts
        for row in range(self.tb_config['data_start_row'], self.tb_sheet.max_row + 1):
            if row not in self.account_mappings:
                account_name = self.tb_sheet.cell(row, self.tb_config['account_name_col']).value
                if account_name:
                    self.tb_sheet.cell(row, hyperlink_col, 'N/A')
        
        print(f"Hyperlinks added in column {get_column_letter(hyperlink_col)}")
        
    def save_workbook(self):
        """Save the modified workbook."""
        print(f"\nSaving linked workbook as: {self.output_file}")
        self.tb_wb.save(self.output_file)
        print("Done! The linked workbook has been created successfully.")
        
    def run(self):
        """Execute the complete linking process."""
        try:
            self.load_workbooks()
            self.analyze_tb_structure()
            self.analyze_gl_structure()
            self.match_accounts()
            self.copy_gl_sheet()
            self.add_hyperlinks()
            self.save_workbook()
            
            print(f"\n✅ Success! Your linked file is ready: {self.output_file}")
            print(f"   - {len(self.account_mappings)} accounts linked")
            print(f"   - GL data copied as 'General Ledger Detail' sheet")
            print(f"   - Click 'View Details' links to navigate to GL accounts")
            
        except Exception as e:
            print(f"\n❌ Error: {str(e)}")
            sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description='Link Trial Balance accounts to General Ledger details',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s TB.xlsx GL.xlsx
  %(prog)s trial_balance.xlsx general_ledger.xlsx -o combined.xlsx
        """
    )
    
    parser.add_argument('tb_file', help='Trial Balance Excel file')
    parser.add_argument('gl_file', help='General Ledger Excel file')
    parser.add_argument('-o', '--output', help='Output file name (default: TB_file_linked.xlsx)')
    
    args = parser.parse_args()
    
    # Validate input files exist
    if not os.path.exists(args.tb_file):
        print(f"Error: TB file '{args.tb_file}' not found")
        sys.exit(1)
    
    if not os.path.exists(args.gl_file):
        print(f"Error: GL file '{args.gl_file}' not found")
        sys.exit(1)
    
    # Create linker and run
    linker = TBGLLinker(args.tb_file, args.gl_file, args.output)
    linker.run()


if __name__ == '__main__':
    main()