#!/usr/bin/env python3
"""
Web-based TB-GL Linker (Net Movement Version) using Streamlit
Links TB accounts to Net Movement figures instead of account headers.
The hyperlink text displays the actual Net Movement value.

Version: 1.1.0 (2026-01-27) - Dynamic hyperlinks with formula support
"""

import streamlit as st
import pandas as pd
import tempfile
import os
try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("Missing required package: openpyxl. Please contact support.")
    st.stop()
from difflib import SequenceMatcher
import re


class TBGLLinkerWeb:
    """TB-GL Linker that links to Net Movement figures instead of account headers."""

    def __init__(self, tb_path, gl_path, output_path):
        self.tb_file = tb_path
        self.gl_file = gl_path
        self.output_file = output_path
        self.tb_wb = None
        self.gl_wb = None
        self.tb_sheet = None
        self.gl_sheet = None
        self.account_mappings = {}
        self.tb_config = {}
        self.gl_accounts = {}  # {account_name: {header_row, net_movement_row, net_movement_col, target_cell, value}}
        self.gl_debit_col = None
        self.gl_credit_col = None

    def load_workbooks(self):
        """Load both workbooks and identify the main sheets."""
        self.tb_wb = load_workbook(self.tb_file)
        self.gl_wb = load_workbook(self.gl_file, data_only=True)

        # Find TB sheet (usually first sheet or one with 'TB' in name)
        tb_sheet_name = self._find_sheet(self.tb_wb, ['TB', 'Trial Balance', 'Trial_Balance'])
        self.tb_sheet = self.tb_wb[tb_sheet_name]

        # Find GL sheet
        gl_sheet_name = self._find_sheet(self.gl_wb, ['GL', 'General Ledger', 'General_Ledger'])
        self.gl_sheet = self.gl_wb[gl_sheet_name]

    def _find_sheet(self, workbook, keywords):
        """Find sheet by keywords in sheet names."""
        sheet_names = workbook.sheetnames

        # First try exact match
        for sheet in sheet_names:
            for keyword in keywords:
                if keyword.lower() in sheet.lower():
                    return sheet

        # If no match found, return first sheet
        return sheet_names[0]

    def analyze_tb_structure(self):
        """Analyze TB structure to find account columns."""
        # Find header row (usually contains 'Account', 'Debit', 'Credit')
        header_row = None
        account_col = None
        account_name_col = None
        debit_col = None
        credit_col = None

        for row in range(1, min(20, self.tb_sheet.max_row + 1)):
            row_values = []
            for col in range(1, min(20, self.tb_sheet.max_column + 1)):
                cell_value = self.tb_sheet.cell(row, col).value
                if cell_value:
                    row_values.append((col, str(cell_value).lower()))

            # Look for key headers
            for col, value in row_values:
                if any(keyword in value for keyword in ['account', 'acct', 'code']):
                    if 'name' in value or 'description' in value:
                        account_name_col = col
                    else:
                        account_col = col
                    header_row = row
                elif 'debit' in value:
                    debit_col = col
                elif 'credit' in value:
                    credit_col = col

            if header_row and (account_col or account_name_col) and (debit_col or credit_col):
                break

        if not header_row:
            raise ValueError("Could not find header row in Trial Balance")

        self.tb_config = {
            'header_row': header_row,
            'account_col': account_col,
            'account_name_col': account_name_col,
            'debit_col': debit_col,
            'credit_col': credit_col,
            'data_start_row': header_row + 1
        }

    def _find_gl_debit_credit_cols(self, near_row=None):
        """Detect Debit/Credit column positions in GL sheet."""
        if self.gl_debit_col and self.gl_credit_col:
            return self.gl_debit_col, self.gl_credit_col

        debit_col = None
        credit_col = None

        # Search in header area first (rows 1-20)
        search_rows = list(range(1, min(21, self.gl_sheet.max_row + 1)))

        # Also search near the specified row if provided
        if near_row:
            search_rows.extend(range(max(1, near_row - 5), min(near_row + 10, self.gl_sheet.max_row + 1)))

        for row in search_rows:
            for col in range(1, min(20, self.gl_sheet.max_column + 1)):
                cell_value = self.gl_sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower().strip()
                    if 'debit' in cell_lower and not debit_col:
                        debit_col = col
                    elif 'credit' in cell_lower and not credit_col:
                        credit_col = col

            if debit_col and credit_col:
                break

        # Fallback defaults if not found (common positions)
        if not debit_col:
            debit_col = 5  # Column E
        if not credit_col:
            credit_col = 6  # Column F

        self.gl_debit_col = debit_col
        self.gl_credit_col = credit_col

        return debit_col, credit_col

    def _get_nonzero_column(self, row, debit_col, credit_col):
        """Get the column with a non-zero value. Returns (column_number, value)."""
        debit_value = self.gl_sheet.cell(row, debit_col).value
        credit_value = self.gl_sheet.cell(row, credit_col).value

        # Convert to numbers, handling None and strings
        try:
            debit_num = float(debit_value) if debit_value else 0
        except (ValueError, TypeError):
            debit_num = 0

        try:
            credit_num = float(credit_value) if credit_value else 0
        except (ValueError, TypeError):
            credit_num = 0

        # If both have values, use Credit (net typically shows there)
        if credit_num != 0:
            return credit_col, credit_num
        elif debit_num != 0:
            return debit_col, debit_num
        else:
            # Both are 0, default to Debit column
            return debit_col, 0

    def _find_net_movement(self, header_row, next_account_row=None):
        """
        Search rows below account header for "Net Movement" text.
        Returns dict with net_movement info or None if not found.
        """
        debit_col, credit_col = self._find_gl_debit_credit_cols(header_row)

        # Determine search limit (stop at next account or reasonable limit)
        max_search = next_account_row if next_account_row else header_row + 500
        max_search = min(max_search, self.gl_sheet.max_row + 1)

        for row in range(header_row + 1, max_search):
            # Check first few columns for "Net Movement" text
            for col in range(1, min(5, self.gl_sheet.max_column + 1)):
                cell_value = self.gl_sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    if 'net movement' in cell_value.lower():
                        # Found Net Movement row - get the non-zero column
                        target_col, value = self._get_nonzero_column(row, debit_col, credit_col)
                        target_cell = f"{get_column_letter(target_col)}{row}"

                        return {
                            'net_movement_row': row,
                            'net_movement_col': target_col,
                            'target_cell': target_cell,
                            'value': value
                        }

        return None

    def _is_account_header(self, row, value):
        """Determine if a cell contains an account header."""
        if row >= self.gl_sheet.max_row:
            return False

        # Account headers usually don't have dates or numbers in the same row
        for col in range(2, min(10, self.gl_sheet.max_column + 1)):
            cell_value = self.gl_sheet.cell(row, col).value
            if cell_value:
                return False

        # Check if following rows have transaction-like data
        for check_row in range(row + 1, min(row + 5, self.gl_sheet.max_row + 1)):
            cell_value = self.gl_sheet.cell(check_row, 1).value
            if cell_value:
                return True

        return False

    def _is_date_like(self, value):
        """Check if string looks like a date."""
        date_patterns = [r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', r'\d{4}[/-]\d{1,2}[/-]\d{1,2}']
        return any(re.match(pattern, value) for pattern in date_patterns)

    def analyze_gl_structure(self):
        """Analyze GL structure to find account sections and their Net Movement rows."""
        # First pass: find all account headers and their row numbers
        account_headers = []

        for row in range(1, self.gl_sheet.max_row + 1):
            # Check first few columns for account names
            for col in range(1, min(5, self.gl_sheet.max_column + 1)):
                cell_value = self.gl_sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    # Check if this looks like an account header
                    if col == 1 and self._is_account_header(row, cell_value):
                        account_name = cell_value.strip()
                        account_headers.append((account_name, row))

        # Detect Debit/Credit columns
        self._find_gl_debit_credit_cols()

        # Second pass: find Net Movement for each account
        for i, (account_name, header_row) in enumerate(account_headers):
            # Determine the next account's row (to limit search)
            next_account_row = account_headers[i + 1][1] if i + 1 < len(account_headers) else None

            # Find Net Movement row
            net_movement_info = self._find_net_movement(header_row, next_account_row)

            if net_movement_info:
                self.gl_accounts[account_name] = {
                    'header_row': header_row,
                    'net_movement_row': net_movement_info['net_movement_row'],
                    'net_movement_col': net_movement_info['net_movement_col'],
                    'target_cell': net_movement_info['target_cell'],
                    'value': net_movement_info['value']
                }
            else:
                # Fallback to account header
                self.gl_accounts[account_name] = {
                    'header_row': header_row,
                    'net_movement_row': header_row,
                    'net_movement_col': 1,
                    'target_cell': f"A{header_row}",
                    'value': None
                }

    def match_accounts(self):
        """Match TB accounts to GL accounts using fuzzy matching."""
        # Get TB accounts
        tb_accounts = []
        for row in range(self.tb_config['data_start_row'], self.tb_sheet.max_row + 1):
            account_name = None

            # Try to get account name from name column first
            if self.tb_config['account_name_col']:
                account_name = self.tb_sheet.cell(row, self.tb_config['account_name_col']).value

            # If no name column or empty, try to find account name in other columns
            if not account_name:
                # Check columns near the account code column for account names
                for col_offset in [1, -1, 2, -2]:  # Check adjacent columns
                    check_col = (self.tb_config['account_col'] or 1) + col_offset
                    if 1 <= check_col <= self.tb_sheet.max_column:
                        cell_value = self.tb_sheet.cell(row, check_col).value
                        if cell_value and isinstance(cell_value, str) and len(cell_value) > 3:
                            # Check if it looks like an account name (not a number or code)
                            if not cell_value.replace('.', '').replace('-', '').isdigit():
                                account_name = cell_value
                                if not self.tb_config['account_name_col']:
                                    self.tb_config['account_name_col'] = check_col
                                break

            if account_name and isinstance(account_name, str):
                tb_accounts.append((row, account_name.strip()))

        # Match each TB account to GL account
        for tb_row, tb_account in tb_accounts:
            best_match = None
            best_score = 0

            for gl_account, gl_info in self.gl_accounts.items():
                # Calculate similarity score
                score = SequenceMatcher(None, tb_account.lower(), gl_account.lower()).ratio()

                # Bonus for exact match
                if tb_account.lower() == gl_account.lower():
                    score = 1.0

                if score > best_score and score > 0.8:  # 80% similarity threshold
                    best_score = score
                    best_match = (gl_account, gl_info)

            if best_match:
                self.account_mappings[tb_row] = best_match

    def copy_gl_sheet(self):
        """Copy GL sheet to TB workbook."""
        # Create new sheet in TB workbook
        gl_sheet_name = 'General Ledger Detail'
        if gl_sheet_name in self.tb_wb.sheetnames:
            del self.tb_wb[gl_sheet_name]

        new_sheet = self.tb_wb.create_sheet(title=gl_sheet_name)

        # Copy all data
        for row in self.gl_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

    def add_hyperlinks(self):
        """Add hyperlinks to TB sheet pointing to Net Movement cells."""
        # Find the best column for hyperlinks (after Credit or last column)
        if self.tb_config['credit_col']:
            hyperlink_col = self.tb_config['credit_col'] + 1
        elif self.tb_config['debit_col']:
            hyperlink_col = self.tb_config['debit_col'] + 1
        else:
            hyperlink_col = self.tb_sheet.max_column + 1

        # Add header - changed to "GL Reference"
        self.tb_sheet.cell(self.tb_config['header_row'], hyperlink_col, 'GL Reference')

        # Add hyperlinks for matched accounts
        for tb_row, (gl_account, gl_info) in self.account_mappings.items():
            cell = self.tb_sheet.cell(tb_row, hyperlink_col)
            target_cell = gl_info['target_cell']
            value = gl_info['value']

            # Format the value for display
            if value is not None and value != 0:
                # Format number with commas, no decimal if whole number
                if isinstance(value, float) and value == int(value):
                    display_value = f"{int(value):,}"
                else:
                    display_value = f"{value:,.2f}" if isinstance(value, float) else str(value)
            else:
                display_value = "0"

            # Create hyperlink formula with dynamic cell reference as display text
            formula = f"=HYPERLINK(\"#'General Ledger Detail'!{target_cell}\", 'General Ledger Detail'!{target_cell})"
            cell.value = formula

        # Add "N/A" for unmatched accounts
        for row in range(self.tb_config['data_start_row'], self.tb_sheet.max_row + 1):
            if row not in self.account_mappings:
                account_name = self.tb_sheet.cell(row, self.tb_config.get('account_name_col', 2)).value
                if account_name:
                    self.tb_sheet.cell(row, hyperlink_col, 'N/A')

    def save_workbook(self):
        """Save the modified workbook."""
        self.tb_wb.save(self.output_file)


def main():
    st.set_page_config(
        page_title="TB-GL Linker",
        page_icon="📊",
        layout="wide"
    )

    st.title("📊 TB-GL Linker")
    st.markdown("**Link your Trial Balance to General Ledger Net Movement figures**")
    st.info("Links to Net Movement cells and displays the actual value as the hyperlink text. Values update dynamically when GL is modified.")
    st.caption("Version 1.1.1 - DEPLOYMENT TEST 27-JAN-2026 14:30")
    st.warning("⚠️ If you see this message, the deployment is working correctly!")

    # Add demo info and GitHub link
    col1, col2, col3 = st.columns([2, 1, 1])
    with col2:
        st.markdown("[![GitHub](https://img.shields.io/badge/GitHub-View%20Code-blue?logo=github)](https://github.com/tuteke2023/excel_referencing)")
    with col3:
        st.markdown("[![Star](https://img.shields.io/github/stars/tuteke2023/excel_referencing?style=social)](https://github.com/tuteke2023/excel_referencing)")

    # Key differences from original
    with st.expander("🔄 How It Works"):
        st.markdown("""
        **This tool:**
        - Links to Net Movement cell (e.g., `F9846`)
        - Displays the actual Net Movement value (e.g., `5,289`)
        - Automatically detects Debit/Credit columns
        - Uses whichever column has the non-zero value
        """)

    # Sample files info
    with st.expander("💡 Need Sample Files?"):
        st.markdown("""
        **Don't have TB and GL files to test?**

        Create sample Excel files with these structures:

        **Trial Balance (TB.xlsx):**
        - Column A: Account Code (200, 400, etc.)
        - Column B: Account Name (Sales, Expenses, etc.)
        - Column C: Account Type (Revenue, Expense, etc.)
        - Column D: Debit amounts
        - Column E: Credit amounts

        **General Ledger (GL.xlsx):**
        - Account sections with account names as headers
        - Transaction details under each account
        - A "Net Movement" row at the end of each account section
        - Date, Description, Debit, Credit columns
        """)

    # Instructions
    with st.expander("📋 How to Use"):
        st.markdown("""
        1. **Upload your Trial Balance Excel file** (should contain Account Names, Debit, Credit columns)
        2. **Upload your General Ledger Excel file** (should contain account sections with "Net Movement" rows)
        3. **Click 'Process Files'** to automatically match accounts and create hyperlinks
        4. **Download the linked file** with hyperlinks pointing to Net Movement values
        """)

    # File upload section
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("📄 Trial Balance File")
        tb_file = st.file_uploader(
            "Upload Trial Balance Excel file",
            type=['xlsx', 'xls'],
            help="Excel file containing your trial balance data"
        )

    with col2:
        st.subheader("📄 General Ledger File")
        gl_file = st.file_uploader(
            "Upload General Ledger Excel file",
            type=['xlsx', 'xls'],
            help="Excel file containing your general ledger data with Net Movement rows"
        )

    # Processing section
    if tb_file and gl_file:
        st.subheader("🔧 Processing Options")

        col1, col2 = st.columns(2)
        with col1:
            similarity_threshold = st.slider(
                "Account Matching Similarity (%)",
                min_value=50,
                max_value=100,
                value=80,
                help="Minimum similarity required to match TB accounts to GL accounts"
            )

        with col2:
            output_filename = st.text_input(
                "Output Filename",
                value="TB_GL_Linked.xlsx",
                help="Name for the output file"
            )

        # Process button
        if st.button("🚀 Process Files", type="primary"):
            try:
                # Save uploaded files to temporary location
                with tempfile.TemporaryDirectory() as tmp_dir:
                    tb_path = os.path.join(tmp_dir, "tb_temp.xlsx")
                    gl_path = os.path.join(tmp_dir, "gl_temp.xlsx")
                    output_path = os.path.join(tmp_dir, output_filename)

                    # Write uploaded files
                    with open(tb_path, 'wb') as f:
                        f.write(tb_file.read())
                    with open(gl_path, 'wb') as f:
                        f.write(gl_file.read())

                    # Create progress bar
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    # Process files
                    status_text.text("🔍 Analyzing file structures...")
                    progress_bar.progress(10)

                    linker = TBGLLinkerWeb(tb_path, gl_path, output_path)

                    status_text.text("📊 Loading workbooks...")
                    progress_bar.progress(25)
                    linker.load_workbooks()

                    status_text.text("🔍 Analyzing TB structure...")
                    progress_bar.progress(40)
                    linker.analyze_tb_structure()

                    status_text.text("🔍 Finding GL accounts and Net Movement rows...")
                    progress_bar.progress(55)
                    linker.analyze_gl_structure()

                    status_text.text("🔗 Matching accounts...")
                    progress_bar.progress(70)
                    linker.match_accounts()

                    status_text.text("📋 Creating linked file with Net Movement references...")
                    progress_bar.progress(85)
                    linker.copy_gl_sheet()
                    linker.add_hyperlinks()
                    linker.save_workbook()

                    progress_bar.progress(100)
                    status_text.text("✅ Processing complete!")

                    # Display results
                    st.success(f"✅ Successfully processed! {len(linker.account_mappings)} accounts linked to Net Movement figures.")

                    # Show matching results
                    if linker.account_mappings:
                        st.subheader("📊 Matching Results")

                        # Create results dataframe
                        results_data = []
                        for tb_row, (gl_account, gl_info) in linker.account_mappings.items():
                            tb_account = linker.tb_sheet.cell(tb_row, linker.tb_config['account_name_col']).value

                            # Format value for display
                            value = gl_info['value']
                            if value is not None and value != 0:
                                if isinstance(value, float) and value == int(value):
                                    formatted_value = f"{int(value):,}"
                                else:
                                    formatted_value = f"{value:,.2f}" if isinstance(value, float) else str(value)
                            else:
                                formatted_value = "0"

                            results_data.append({
                                'TB Account': tb_account,
                                'GL Account': gl_account,
                                'Target Cell': gl_info['target_cell'],
                                'Net Movement': formatted_value,
                                'Status': '✅ Matched'
                            })

                        df = pd.DataFrame(results_data)
                        st.dataframe(df, use_container_width=True)

                        # Match statistics
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Total Matched", len(results_data))
                        with col2:
                            st.metric("GL Accounts Found", len(linker.gl_accounts))
                        with col3:
                            # Count accounts with Net Movement found
                            net_movement_found = sum(1 for info in linker.gl_accounts.values()
                                                    if info['target_cell'] != f"A{info['header_row']}")
                            st.metric("Net Movement Found", net_movement_found)
                        with col4:
                            match_rate = len(linker.account_mappings) / len(results_data) * 100 if results_data else 0
                            st.metric("Match Rate", f"{match_rate:.1f}%")

                        # Debug info - show sample hyperlinks
                        with st.expander("🔍 Debug: Sample Hyperlinks (click to expand)"):
                            st.write("First 5 matched accounts and their target cells:")
                            for i, (tb_row, (gl_account, gl_info)) in enumerate(list(linker.account_mappings.items())[:5]):
                                is_fallback = gl_info['target_cell'].startswith('A')
                                status = "⚠️ FALLBACK" if is_fallback else "✅ NET MOVEMENT"
                                st.write(f"{status}: {gl_account} → {gl_info['target_cell']} = {gl_info['value']}")

                    # Download button
                    with open(output_path, 'rb') as f:
                        st.download_button(
                            label="📥 Download Linked File",
                            data=f.read(),
                            file_name=output_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

            except Exception as e:
                st.error(f"❌ Error processing files: {str(e)}")
                st.error("Please check your file formats and try again.")


if __name__ == "__main__":
    main()
