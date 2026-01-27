#!/usr/bin/env python3
"""
Excel to Text Converter for Claude Analysis

Converts openpyxl workbook/sheet data to text format suitable for
Claude Code headless analysis.
"""

from openpyxl.utils import get_column_letter


class ExcelToText:
    """Convert Excel data to text format for Claude analysis."""

    @staticmethod
    def sheet_to_csv_preview(sheet, max_rows=50, max_cols=15) -> str:
        """
        Convert openpyxl sheet to CSV-like text for Claude analysis.

        Args:
            sheet: openpyxl worksheet object
            max_rows: Maximum rows to include (default 50)
            max_cols: Maximum columns to include (default 15)

        Returns:
            CSV-formatted string with row numbers
        """
        lines = []

        # Determine actual bounds
        actual_rows = min(max_rows, sheet.max_row)
        actual_cols = min(max_cols, sheet.max_column)

        # Add header with column letters
        header = "Row," + ",".join(get_column_letter(c) for c in range(1, actual_cols + 1))
        lines.append(header)

        # Add data rows
        for row in range(1, actual_rows + 1):
            row_values = [str(row)]
            for col in range(1, actual_cols + 1):
                cell_value = sheet.cell(row, col).value
                if cell_value is None:
                    row_values.append("")
                elif isinstance(cell_value, str):
                    # Escape commas and quotes for CSV format
                    escaped = cell_value.replace('"', '""')
                    if ',' in escaped or '"' in escaped or '\n' in escaped:
                        row_values.append(f'"{escaped}"')
                    else:
                        row_values.append(escaped)
                else:
                    row_values.append(str(cell_value))
            lines.append(",".join(row_values))

        # Add note if truncated
        if sheet.max_row > max_rows:
            lines.append(f"... ({sheet.max_row - max_rows} more rows)")
        if sheet.max_column > max_cols:
            lines.append(f"... ({sheet.max_column - max_cols} more columns)")

        return "\n".join(lines)

    @staticmethod
    def sheet_names_summary(workbook) -> str:
        """
        List sheet names with row/column counts.

        Args:
            workbook: openpyxl workbook object

        Returns:
            Summary string of all sheets
        """
        lines = ["Sheets in workbook:"]
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            lines.append(f"  - '{sheet_name}': {sheet.max_row} rows x {sheet.max_column} cols")
        return "\n".join(lines)

    @staticmethod
    def sample_account_sections(sheet, sample_size=10) -> str:
        """
        Extract potential account header rows with surrounding context.
        Useful for understanding GL structure.

        Args:
            sheet: openpyxl worksheet object
            sample_size: Number of potential sections to sample

        Returns:
            Text showing potential account sections
        """
        sections = []
        current_section = None
        section_count = 0

        for row in range(1, min(sheet.max_row + 1, 500)):  # Limit scan
            col_a_value = sheet.cell(row, 1).value

            # Check if this looks like an account header (text in col A, empty cols B-D)
            if col_a_value and isinstance(col_a_value, str) and len(col_a_value.strip()) > 2:
                # Check if adjacent cells are empty (typical of headers)
                cols_empty = all(
                    sheet.cell(row, c).value is None
                    for c in range(2, min(5, sheet.max_column + 1))
                )

                if cols_empty:
                    if section_count >= sample_size:
                        break

                    # Start new section
                    current_section = {
                        'header': col_a_value.strip(),
                        'header_row': row,
                        'sample_rows': []
                    }
                    sections.append(current_section)
                    section_count += 1

            # Collect a few rows after header
            elif current_section and len(current_section['sample_rows']) < 8:
                row_data = []
                for col in range(1, min(8, sheet.max_column + 1)):
                    val = sheet.cell(row, col).value
                    row_data.append(str(val) if val is not None else "")
                current_section['sample_rows'].append((row, row_data))

        # Format output
        lines = [f"Sample account sections (first {len(sections)}):"]
        for section in sections:
            lines.append(f"\n--- Account: '{section['header']}' (Row {section['header_row']}) ---")
            for row_num, row_data in section['sample_rows']:
                # Format row showing column letters
                formatted = " | ".join(f"{get_column_letter(i+1)}:{v}" for i, v in enumerate(row_data) if v)
                if formatted:
                    lines.append(f"  Row {row_num}: {formatted}")

        return "\n".join(lines)

    @staticmethod
    def get_row_range(sheet, start_row, end_row, max_cols=10) -> str:
        """
        Get a specific range of rows as text.

        Args:
            sheet: openpyxl worksheet object
            start_row: First row to include
            end_row: Last row to include
            max_cols: Maximum columns to include

        Returns:
            Text representation of the row range
        """
        lines = []
        actual_cols = min(max_cols, sheet.max_column)

        for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
            row_values = []
            for col in range(1, actual_cols + 1):
                val = sheet.cell(row, col).value
                if val is not None:
                    row_values.append(f"{get_column_letter(col)}:{val}")
            if row_values:
                lines.append(f"Row {row}: {' | '.join(row_values)}")
            else:
                lines.append(f"Row {row}: (empty)")

        return "\n".join(lines)

    @staticmethod
    def find_text_occurrences(sheet, search_text, max_results=20) -> str:
        """
        Find all occurrences of text in a sheet.

        Args:
            sheet: openpyxl worksheet object
            search_text: Text to search for (case-insensitive)
            max_results: Maximum results to return

        Returns:
            Text listing all occurrences with cell references
        """
        results = []
        search_lower = search_text.lower()

        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                val = sheet.cell(row, col).value
                if val and isinstance(val, str) and search_lower in val.lower():
                    cell_ref = f"{get_column_letter(col)}{row}"
                    results.append(f"  {cell_ref}: '{val}'")
                    if len(results) >= max_results:
                        break
            if len(results) >= max_results:
                break

        if results:
            return f"Found '{search_text}' in:\n" + "\n".join(results)
        else:
            return f"'{search_text}' not found in sheet"
