#!/usr/bin/env python3
"""
Quick TB-GL Linker - Simplified version for basic use cases
"""

import openpyxl
import sys

def link_tb_gl(tb_file='TB.xlsx', gl_file='GL.xlsx'):
    """Quick link function with minimal configuration."""
    print(f"Linking {tb_file} with {gl_file}...")
    
    # Load workbooks
    tb_wb = openpyxl.load_workbook(tb_file)
    gl_wb = openpyxl.load_workbook(gl_file)
    
    # Get first sheets
    tb_sheet = tb_wb.active
    gl_sheet = gl_wb.active
    
    # Copy GL sheet to TB workbook
    if 'General Ledger Detail' in tb_wb.sheetnames:
        del tb_wb['General Ledger Detail']
    
    new_sheet = tb_wb.create_sheet('General Ledger Detail')
    for row in gl_sheet.iter_rows(values_only=True):
        new_sheet.append(row)
    
    # Find account column in TB (usually column B)
    account_col = 2  # Default to column B
    
    # Add hyperlinks
    max_col = tb_sheet.max_column
    link_col = max_col + 1
    
    tb_sheet.cell(1, link_col, 'Reference')
    
    # Simple matching - look for account names in GL
    for tb_row in range(2, tb_sheet.max_row + 1):
        account_name = tb_sheet.cell(tb_row, account_col).value
        if account_name:
            # Search for this account in GL
            for gl_row in range(1, new_sheet.max_row + 1):
                gl_value = new_sheet.cell(gl_row, 1).value
                if gl_value and str(account_name).lower() in str(gl_value).lower():
                    formula = f'=HYPERLINK("#\'General Ledger Detail\'!A{gl_row}", "View")'
                    tb_sheet.cell(tb_row, link_col, formula)
                    break
    
    # Save
    output_file = tb_file.replace('.xlsx', '_linked.xlsx')
    tb_wb.save(output_file)
    print(f"✅ Saved as: {output_file}")

if __name__ == '__main__':
    if len(sys.argv) == 3:
        link_tb_gl(sys.argv[1], sys.argv[2])
    else:
        link_tb_gl()