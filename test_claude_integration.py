#!/usr/bin/env python3
"""
Test script for Claude Code headless integration.
Creates sample Excel files and tests the analysis pipeline.
"""

import tempfile
import os
from openpyxl import Workbook
from excel_converter import ExcelToText
from claude_analyzer import ClaudeAnalyzer, ClaudeAnalyzerWithFallback


def create_sample_tb():
    """Create a sample Trial Balance workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    # Headers
    ws['A1'] = "Account Code"
    ws['B1'] = "Account Name"
    ws['C1'] = "Debit"
    ws['D1'] = "Credit"

    # Sample data
    data = [
        ("1000", "Cash at Bank", 50000, 0),
        ("1100", "Accounts Receivable", 25000, 0),
        ("2000", "Accounts Payable", 0, 15000),
        ("3000", "Share Capital", 0, 100000),
        ("4000", "Sales Revenue", 0, 75000),
        ("5000", "Cost of Goods Sold", 45000, 0),
        ("6000", "Operating Expenses", 70000, 0),
    ]

    for i, (code, name, debit, credit) in enumerate(data, start=2):
        ws[f'A{i}'] = code
        ws[f'B{i}'] = name
        ws[f'C{i}'] = debit if debit else None
        ws[f'D{i}'] = credit if credit else None

    return wb


def create_sample_gl():
    """Create a sample General Ledger workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "General Ledger"

    row = 1

    accounts = [
        ("Cash at Bank", [
            ("2026-01-01", "Opening Balance", 40000, 0),
            ("2026-01-15", "Customer Payment", 15000, 0),
            ("2026-01-20", "Supplier Payment", 0, 5000),
        ], 50000),
        ("Accounts Receivable", [
            ("2026-01-05", "Invoice #001", 30000, 0),
            ("2026-01-15", "Payment Received", 0, 5000),
        ], 25000),
        ("Sales Revenue", [
            ("2026-01-05", "Invoice #001", 0, 30000),
            ("2026-01-10", "Invoice #002", 0, 45000),
        ], 75000),
    ]

    for account_name, transactions, net_movement in accounts:
        # Account header
        ws[f'A{row}'] = account_name
        row += 1

        # Column headers
        ws[f'A{row}'] = "Date"
        ws[f'B{row}'] = "Description"
        ws[f'C{row}'] = "Debit"
        ws[f'D{row}'] = "Credit"
        row += 1

        # Transactions
        for date, desc, debit, credit in transactions:
            ws[f'A{row}'] = date
            ws[f'B{row}'] = desc
            ws[f'C{row}'] = debit if debit else None
            ws[f'D{row}'] = credit if credit else None
            row += 1

        # Net Movement row
        ws[f'A{row}'] = "Net Movement"
        ws[f'D{row}'] = net_movement
        row += 2  # Blank row between accounts

    return wb


def test_excel_converter():
    """Test ExcelToText conversion."""
    print("\n=== Testing ExcelToText ===")

    tb_wb = create_sample_tb()
    tb_sheet = tb_wb.active

    print("\n1. Sheet to CSV preview:")
    preview = ExcelToText.sheet_to_csv_preview(tb_sheet, max_rows=10)
    print(preview[:500])

    print("\n2. Sheet names summary:")
    summary = ExcelToText.sheet_names_summary(tb_wb)
    print(summary)

    gl_wb = create_sample_gl()
    gl_sheet = gl_wb.active

    print("\n3. Sample account sections (GL):")
    sections = ExcelToText.sample_account_sections(gl_sheet, sample_size=3)
    print(sections[:800])

    return True


def test_claude_analyzer():
    """Test Claude analyzer (if available)."""
    print("\n=== Testing ClaudeAnalyzer ===")

    analyzer = ClaudeAnalyzer(timeout=60)

    if not analyzer.is_available():
        print("Claude CLI not available, skipping Claude tests")
        return True

    print("Claude CLI is available!")

    # Create sample workbooks
    tb_wb = create_sample_tb()
    gl_wb = create_sample_gl()

    # Test sheet identification
    print("\n1. Testing sheet identification...")
    result = analyzer.identify_sheets(tb_wb, gl_wb)
    if result:
        print(f"   Result: {result}")
    else:
        print("   No result (may need fallback)")

    # Test TB structure analysis
    print("\n2. Testing TB structure analysis...")
    result = analyzer.analyze_tb_structure(tb_wb.active)
    if result:
        print(f"   Result: {result}")
    else:
        print("   No result (may need fallback)")

    # Test GL structure analysis
    print("\n3. Testing GL structure analysis...")
    result = analyzer.analyze_gl_structure(gl_wb.active)
    if result:
        print(f"   Result: {result}")
    else:
        print("   No result (may need fallback)")

    return True


def test_web_app_integration():
    """Test the web app integration."""
    print("\n=== Testing Web App Integration ===")

    from web_app import TBGLLinkerWeb, CLAUDE_AVAILABLE
    print(f"CLAUDE_AVAILABLE: {CLAUDE_AVAILABLE}")

    # Create temp files
    with tempfile.TemporaryDirectory() as tmp_dir:
        tb_path = os.path.join(tmp_dir, "tb.xlsx")
        gl_path = os.path.join(tmp_dir, "gl.xlsx")
        output_path = os.path.join(tmp_dir, "output.xlsx")

        # Save sample workbooks
        tb_wb = create_sample_tb()
        gl_wb = create_sample_gl()
        tb_wb.save(tb_path)
        gl_wb.save(gl_path)

        # Test with Claude enabled
        print("\n1. Testing with Claude enabled...")
        linker = TBGLLinkerWeb(tb_path, gl_path, output_path, use_claude=True)
        linker.load_workbooks()
        linker.analyze_tb_structure()
        linker.analyze_gl_structure()
        linker.match_accounts()

        print(f"   TB config: {linker.tb_config}")
        print(f"   GL accounts found: {len(linker.gl_accounts)}")
        print(f"   Matches: {len(linker.account_mappings)}")
        print(f"   Claude used: {linker.claude_used}")

        # Test with Claude disabled (fallback)
        print("\n2. Testing with Claude disabled (fallback)...")
        linker2 = TBGLLinkerWeb(tb_path, gl_path, output_path, use_claude=False)
        linker2.load_workbooks()
        linker2.analyze_tb_structure()
        linker2.analyze_gl_structure()
        linker2.match_accounts()

        print(f"   TB config: {linker2.tb_config}")
        print(f"   GL accounts found: {len(linker2.gl_accounts)}")
        print(f"   Matches: {len(linker2.account_mappings)}")
        print(f"   Claude used: {linker2.claude_used}")

    return True


if __name__ == "__main__":
    print("=" * 60)
    print("TB-GL Linker Claude Integration Test")
    print("=" * 60)

    try:
        test_excel_converter()
        test_claude_analyzer()
        test_web_app_integration()
        print("\n" + "=" * 60)
        print("All tests completed successfully!")
        print("=" * 60)
    except Exception as e:
        print(f"\nTest failed with error: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
